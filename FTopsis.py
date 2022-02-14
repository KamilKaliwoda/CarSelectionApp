import numpy as np
import openpyxl as op
import pandas as pd
import operator


# Klasa odpowiadająca za każdy punkt
class Point:
    id = 1

    def __init__(self, cirts, name):
        self.name = name
        self.class_number = 0
        self.criteria = cirts
        self.d_plus = 0
        self.d_minus = 0
        self.score = 0
        self.id = Point.id
        Point.id = Point.id + 1


# Wyznaczenie punktów idealnego i antyidealnego
def imaginary(set_of_pts):

    p_ideal = list()
    p_antyideal = list()

    for criterium in range(len(set_of_pts[0].criteria)):

        minimum = np.inf
        maximum = 0

        for point in set_of_pts:
            if point.criteria[criterium] < minimum:
                minimum = point.criteria[criterium]

            elif point.criteria[criterium] > maximum:
                maximum = point.criteria[criterium]

        p_ideal.append(minimum)
        p_antyideal.append(maximum)

    p_ideal = np.array(p_ideal)
    p_antyideal = np.array(p_antyideal)

    return p_ideal, p_antyideal


# Obliczanie odleglosci pomiedzy punktami
def distance(u_point, to_point):
    dist = 0
    for i in range(len(u_point)):
        dist = dist + ((to_point[i] - u_point[i])**2)

    dist = np.sqrt(dist)

    return dist


# Zmienne i kryteria
# Prioritize_disesel = False
cols_to_maximize = ["Rocznik", "Moc silnika", "Pojemność silnika", "Ilość drzwi", "Ilość miejsc",
                    "Pojemność bagażnika", "Klimatyzacja", "Gwarancja", "Opinia"]


def load_database(criteria, Prioritize_diesel = False):
    # Wczytanie z arkusza do dataframe'a
    wb_obj = op.load_workbook("Baza_samochodow.xlsx")
    ws = wb_obj['Arkusz1']
    data = ws.values
    index = next(data)[0:]
    df = pd.DataFrame(data, columns=index)

    # Zamiana stringów na wartości 
    df["Skrzynia biegów"].replace({"Automatyczna" : 0, "Manualna" : 1}, inplace=True)
    if Prioritize_diesel:
        df["Paliwo"].replace({"Diesel" : 0, "Benzyna" : 1}, inplace=True)
    else:
        df["Paliwo"].replace({"Diesel" : 1, "Benzyna" : 0}, inplace=True)

    # Normalizacja i odwrócenie jeżeli maksymalizujemy
    columns = df.columns
    for col in columns:
        if col == "Marka i Model":
            continue
        norm = np.amax(df[col].values)
        if col in cols_to_maximize:
            normalized = 1 - (df[col].values / norm)
        else:
            normalized = df[col].values / norm
        df[col] = normalized

    # Wyłuskanie 3 kryteriów jeśli żądane
    if len(criteria) == 0:
        index = df.columns
        index = index[1:]
        df = df.values
        
    else:
        criteria_with_name = ["Marka i Model"] + criteria
        index = df[criteria].columns
        df = df[criteria_with_name].values

    return df, index


# Funkcja tworzaca ranking metoda Fuzy Topsis
def fuzzy_topsis(crits, Prioritize_diesel):
    [database, index] = load_database(crits, Prioritize_diesel)

    set_of_points = list()
    for row in range(len(database)):
        set_of_points.append(Point(database[row][1:],database[row, 0]))

    ideal, antyideal = imaginary(set_of_points)

    for point in set_of_points:
        point.d_plus = distance(point.criteria, ideal)
        point.d_minus = distance(point.criteria, antyideal)
        point.score = point.d_minus/(point.d_minus + point.d_plus)

    sorted_set_of_points = sorted(set_of_points, key=operator.attrgetter('score'), reverse=True)
    Point.id = 1

    return sorted_set_of_points


# Konwersja wyniku do formy prezentowalnej w GUI
def ranking_to_dataframe_all(n_rows, Prioritize_diesel = False):
    criteria = []
    ft_result = fuzzy_topsis(criteria, Prioritize_diesel)
    wb_obj = op.load_workbook("Baza_samochodow.xlsx")
    ws = wb_obj['Arkusz1']
    data = ws.values
    index = next(data)[0:]
    df = pd.DataFrame(data, columns=index)

    score_list = []
    for index in range(len(df) + 1):
        for car in ft_result:
            if car.id == index:
                score_list.append(car.score)

    df["Ranking"] = score_list

    df_result = df.sort_values(by=["Ranking"], ascending=False)

    return df_result.head(n_rows)


def ranking_to_dataframe_three(crits, n_rows): # -> tuple[pd.DataFrame, list[list[list]]]:
    ft_result = fuzzy_topsis(crits, Prioritize_diesel = False)
    wb_obj = op.load_workbook("Baza_samochodow.xlsx")
    ws = wb_obj['Arkusz1']
    data = ws.values
    index = next(data)[0:]
    df = pd.DataFrame(data, columns=index)

    score_list = []
    for index in range(len(df) + 1):
        for car in ft_result:
            if car.id == index:
                score_list.append(car.score)

    df["Ranking"] = score_list

    df_result = df.sort_values(by=["Ranking"], ascending=False)
    for col in df.columns:
        if col == "Marka i Model":
            continue
        if col == "Ranking":
            continue
        if col not in crits:
            df_result = df_result.drop(col, 1)

    
    df_result = df_result.head(n_rows)
    df_result2 = pd.DataFrame() 
    df_result2['Marka i Model'] = df_result["Marka i Model"]
    df_result2[crits[0]] = df_result[crits[0]]
    df_result2[crits[1]] = df_result[crits[1]]
    df_result2[crits[2]] = df_result[crits[2]]
    
    return df_result2, [df_result2.drop(df_result2.columns[[0]], axis=1).values.tolist()] # [df_result2.drop(df_result2.columns[[0, 4]], axis=1).values.tolist()]



# result_df = ranking_to_dataframe_all(99, True)
# result_df.to_excel("Ftopsis.xlsx")

#criteria = ["Rocznik", "Moc silnika", "Pojemność bagażnika"]
#result_df = ranking_to_dataframe_three(criteria, 6)
#print(result_df)
